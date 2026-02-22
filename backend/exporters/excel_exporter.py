"""
Excel Exporter — All 21 Document Types

Exports extraction results to Excel with type-specific formatted sheets.
"""

from pathlib import Path
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from logger import logger
import config


class ExcelExporter:
    """Export structured data to Excel with formatting"""

    def __init__(self):
        self.header_fill   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font   = Font(bold=True, color="FFFFFF")
        self.section_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font  = Font(bold=True)
        self.alt_fill      = PatternFill(start_color="EBF1FB", end_color="EBF1FB", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Map document types → specialized sheet builder
        self._sheet_builders = {
            'invoice':           self._sheet_invoice,
            'resume':            self._sheet_resume,
            'contract':          self._sheet_contract,
            'prescription':      self._sheet_prescription,
            'certificate':       self._sheet_certificate,
            'bank_statement':    self._sheet_bank_statement,
            'utility_bill':      self._sheet_utility_bill,
            'marksheet':         self._sheet_marksheet,
            'ration_card':       self._sheet_ration_card,
            'aadhar_card':       self._sheet_aadhar_card,
            'pan_card':          self._sheet_pan_card,
            'driving_licence':   self._sheet_driving_licence,
            'passport':          self._sheet_passport,
            'cheque':            self._sheet_cheque,
            'form_16':           self._sheet_form_16,
            'insurance_policy':  self._sheet_insurance_policy,
            'gst_certificate':   self._sheet_gst_certificate,
            'birth_certificate': self._sheet_birth_certificate,
            'death_certificate': self._sheet_death_certificate,
            'land_record':       self._sheet_land_record,
            'nrega_card':        self._sheet_nrega_card,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # Public API
    # ─────────────────────────────────────────────────────────────────────────

    def export(self, extraction_result: Dict[str, Any], output_path: str = None) -> str:
        if not output_path:
            output_path = config.get_output_filename("extracted", ".xlsx")
        output_path = Path(output_path)

        wb = Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(wb, extraction_result)

        if 'structured_data' in extraction_result:
            doc_type = extraction_result.get('classification', {}).get('document_type', 'unknown')
            builder = self._sheet_builders.get(doc_type)
            if builder:
                builder(wb, extraction_result)

        if 'comprehensive_confidence' in extraction_result:
            self._create_confidence_sheet(wb, extraction_result)

        wb.save(output_path)
        logger.info(f"Exported to Excel: {output_path.name}")
        return str(output_path)

    # ─────────────────────────────────────────────────────────────────────────
    # Helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _hrow(self, ws, row: int, headers: list):
        """Write formatted header row."""
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = self.header_font
            c.fill = self.header_fill
            c.alignment = Alignment(horizontal='center')

    def _kv(self, ws, row: int, key: str, value) -> int:
        """Write a key-value pair and return next row."""
        ws[f'A{row}'] = key
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = str(value) if value is not None else ''
        return row + 1

    def _section(self, ws, row: int, title: str) -> int:
        """Write a bold section header spanning A:D and return next row."""
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        return row + 1

    def _dict_rows(self, ws, row: int, d: dict, prefix: str = '') -> int:
        """Flatten a dict and write each key-value."""
        for k, v in (d or {}).items():
            label = f"{prefix}{k.replace('_', ' ').title()}"
            if isinstance(v, dict):
                row = self._dict_rows(ws, row, v, prefix=label + ' → ')
            elif isinstance(v, list):
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = ', '.join(str(i) for i in v) if v else ''
                row += 1
            else:
                row = self._kv(ws, row, label, v)
        return row

    def _set_col_widths(self, ws, widths: list):
        """Set column widths from list [col_A_width, col_B_width, ...]."""
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w

    # ─────────────────────────────────────────────────────────────────────────
    # Summary + Confidence sheets (for all types)
    # ─────────────────────────────────────────────────────────────────────────

    def _create_summary_sheet(self, wb, result):
        ws = wb.create_sheet("Summary")
        ws['A1'] = "ADIVA — Extraction Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        row = 3

        # Metadata
        row = self._section(ws, row, "Document Information")
        for k, v in result.get('metadata', {}).items():
            if not isinstance(v, (dict, list)):
                row = self._kv(ws, row, k.replace('_', ' ').title(), v)
        row += 1

        # Classification
        if 'classification' in result:
            row = self._section(ws, row, "AI Classification")
            cls = result['classification']
            row = self._kv(ws, row, "Document Type", cls.get('document_type', ''))
            row = self._kv(ws, row, "Confidence", f"{cls.get('confidence', 0) * 100:.1f}%")
            row = self._kv(ws, row, "Reasoning", cls.get('reasoning', ''))
            row = self._kv(ws, row, "Alternative Type", cls.get('alternative_type', ''))

        self._set_col_widths(ws, [28, 50, 20])

    def _create_confidence_sheet(self, wb, result):
        ws = wb.create_sheet("Confidence Metrics")
        conf = result['comprehensive_confidence']
        row = 1
        ws[f'A{row}'] = "Extraction Quality Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        row = self._kv(ws, row, "Overall Confidence", f"{conf['overall_confidence'] * 100:.1f}%")
        ws[f'B{row-1}'].font = Font(bold=True, size=12)
        row = self._kv(ws, row, "Grade", conf['grade'])
        ws[f'B{row-1}'].font = Font(bold=True, size=12, color="276221" if conf['grade'] in ('A+','A','A-') else "C55A11")
        row += 1
        row = self._section(ws, row, "Detailed Metrics")
        for metric, value in conf['metrics'].items():
            row = self._kv(ws, row, metric.replace('_', ' ').title(), f"{value * 100:.1f}%")
        row += 1
        row = self._section(ws, row, "Explanations")
        for exp in conf.get('explanations', []):
            ws[f'A{row}'] = exp; row += 1
        row += 1
        row = self._section(ws, row, "Recommendations")
        for rec in conf.get('recommendations', []):
            ws[f'A{row}'] = rec; row += 1
        self._set_col_widths(ws, [35, 25])

    # ─────────────────────────────────────────────────────────────────────────
    # General Documents
    # ─────────────────────────────────────────────────────────────────────────

    def _sheet_invoice(self, wb, result):
        ws = wb.create_sheet("Invoice Data")
        d = result['structured_data']
        row = self._section(ws, 1, "Invoice Details")
        row = self._kv(ws, row, "Invoice Number", d.get('invoice_number'))
        row = self._kv(ws, row, "Invoice Date", d.get('invoice_date'))
        row = self._kv(ws, row, "Due Date", d.get('due_date'))
        row = self._kv(ws, row, "Currency", d.get('currency'))
        row = self._kv(ws, row, "Payment Terms", d.get('payment_terms'))
        row += 1
        row = self._section(ws, row, "Vendor")
        row = self._dict_rows(ws, row, d.get('vendor', {}))
        row += 1
        row = self._section(ws, row, "Customer")
        row = self._dict_rows(ws, row, d.get('customer', {}))
        row += 1
        row = self._section(ws, row, "Line Items")
        self._hrow(ws, row, ['Description', 'Qty', 'Unit Price', 'Total'])
        row += 1
        for item in d.get('line_items', []):
            ws[f'A{row}'] = item.get('description', '')
            ws[f'B{row}'] = item.get('quantity', '')
            ws[f'C{row}'] = item.get('unit_price', '')
            ws[f'D{row}'] = item.get('total', '')
            row += 1
        row += 1
        self._kv(ws, row, "Subtotal", d.get('subtotal')); row += 1
        self._kv(ws, row, "Tax", d.get('tax')); row += 1
        ws[f'A{row}'] = "TOTAL"; ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = d.get('total', ''); ws[f'B{row}'].font = Font(bold=True, size=12)
        self._set_col_widths(ws, [40, 12, 15, 15])

    def _sheet_resume(self, wb, result):
        ws = wb.create_sheet("Resume Data")
        d = result['structured_data']
        row = self._section(ws, 1, "Personal Information")
        row = self._dict_rows(ws, row, d.get('personal_info', {}))
        row += 1
        row = self._section(ws, row, "Work Experience")
        self._hrow(ws, row, ['Company', 'Position', 'Location', 'Duration'])
        row += 1
        for exp in d.get('experience', []):
            ws[f'A{row}'] = exp.get('company', '')
            ws[f'B{row}'] = exp.get('position', '')
            ws[f'C{row}'] = exp.get('location', '')
            ws[f'D{row}'] = exp.get('duration', exp.get('start_date',''))
            row += 1
        row += 1
        row = self._section(ws, row, "Education")
        self._hrow(ws, row, ['Degree', 'Institution', 'Year', 'GPA/Grade'])
        row += 1
        for edu in d.get('education', []):
            ws[f'A{row}'] = edu.get('degree', '')
            ws[f'B{row}'] = edu.get('institution', '')
            ws[f'C{row}'] = edu.get('year', '')
            ws[f'D{row}'] = edu.get('gpa', '')
            row += 1
        row += 1
        row = self._section(ws, row, "Skills")
        for s in d.get('skills', []):
            ws[f'A{row}'] = s; row += 1
        self._set_col_widths(ws, [30, 30, 20, 20])

    def _sheet_contract(self, wb, result):
        ws = wb.create_sheet("Contract Data")
        d = result['structured_data']
        row = self._section(ws, 1, "Contract Details")
        for key in ['contract_title','contract_type','contract_number','contract_date',
                    'effective_date','expiration_date','term_duration','governing_law']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Parties")
        self._hrow(ws, row, ['Name', 'Role', 'Address', 'Signatory'])
        row += 1
        for p in d.get('parties', []):
            ws[f'A{row}'] = p.get('name','')
            ws[f'B{row}'] = p.get('role','')
            ws[f'C{row}'] = p.get('address','')
            ws[f'D{row}'] = p.get('signatory','')
            row += 1
        row += 1
        row = self._section(ws, row, "Key Clauses")
        row = self._kv(ws, row, "Termination", d.get('termination_clause'))
        row = self._kv(ws, row, "Confidentiality", d.get('confidentiality'))
        row = self._kv(ws, row, "Dispute Resolution", d.get('dispute_resolution'))
        self._set_col_widths(ws, [28, 20, 35, 25])

    def _sheet_prescription(self, wb, result):
        ws = wb.create_sheet("Prescription Data")
        d = result['structured_data']
        row = self._section(ws, 1, "Doctor Information")
        row = self._dict_rows(ws, row, d.get('doctor', {}))
        row += 1
        row = self._section(ws, row, "Patient Information")
        row = self._dict_rows(ws, row, d.get('patient', {}))
        row += 1
        row = self._section(ws, row, "Medicines")
        self._hrow(ws, row, ['Medicine', 'Dosage', 'Frequency', 'Duration', 'Instructions'])
        row += 1
        for med in d.get('medicines', []):
            ws[f'A{row}'] = med.get('name','')
            ws[f'B{row}'] = med.get('dosage','')
            ws[f'C{row}'] = med.get('frequency','')
            ws[f'D{row}'] = med.get('duration','')
            ws[f'E{row}'] = med.get('instructions','')
            row += 1
        row += 1
        row = self._kv(ws, row, "Diagnosis", d.get('diagnosis'))
        row = self._kv(ws, row, "Date", d.get('prescription_date'))
        row = self._kv(ws, row, "Follow-up", d.get('follow_up'))
        self._set_col_widths(ws, [28, 18, 18, 15, 30])

    def _sheet_certificate(self, wb, result):
        ws = wb.create_sheet("Certificate Data")
        d = result['structured_data']
        row = self._section(ws, 1, "Certificate Details")
        for key in ['certificate_type','certificate_number','issue_date','event_date','event_place']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Primary Person")
        row = self._dict_rows(ws, row, d.get('primary_person', {}))
        row += 1
        if d.get('secondary_person'):
            row = self._section(ws, row, "Secondary Person")
            row = self._dict_rows(ws, row, d.get('secondary_person', {}))
            row += 1
        row = self._section(ws, row, "Issuing Authority")
        row = self._kv(ws, row, "Authority", d.get('issuing_authority'))
        row = self._kv(ws, row, "Registrar", d.get('registrar_name'))
        if d.get('witnesses'):
            row += 1
            row = self._section(ws, row, "Witnesses")
            for w in d.get('witnesses', []):
                ws[f'A{row}'] = w; row += 1
        self._set_col_widths(ws, [28, 45])

    # ─────────────────────────────────────────────────────────────────────────
    # Civic / Government Documents
    # ─────────────────────────────────────────────────────────────────────────

    def _sheet_bank_statement(self, wb, result):
        ws = wb.create_sheet("Bank Statement")
        d = result['structured_data']
        row = self._section(ws, 1, "Account Information")
        for key in ['bank_name','branch_name','account_number','account_type',
                    'ifsc_code','account_holder','statement_period_from','statement_period_to',
                    'opening_balance','closing_balance']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Transactions")
        self._hrow(ws, row, ['Date', 'Description', 'Debit (₹)', 'Credit (₹)', 'Balance (₹)'])
        row += 1
        for i, txn in enumerate(d.get('transactions', [])):
            fill = self.alt_fill if i % 2 == 0 else None
            for col, key in enumerate(['date','description','debit_amount','credit_amount','balance'], 1):
                c = ws.cell(row=row, column=col, value=txn.get(key, ''))
                if fill: c.fill = fill
            row += 1
        row += 1
        row = self._kv(ws, row, "Total Credits", d.get('total_credits'))
        row = self._kv(ws, row, "Total Debits", d.get('total_debits'))
        self._set_col_widths(ws, [16, 40, 16, 16, 18])

    def _sheet_utility_bill(self, wb, result):
        ws = wb.create_sheet("Utility Bill")
        d = result['structured_data']
        row = self._section(ws, 1, "Consumer Details")
        for key in ['bill_type','provider_name','consumer_name','consumer_number',
                    'meter_number','billing_address']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Billing Details")
        for key in ['billing_period','bill_date','due_date','previous_reading',
                    'current_reading','units_consumed']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Charges")
        row = self._dict_rows(ws, row, d.get('charges', {}))
        row += 1
        ws[f'A{row}'] = "TOTAL AMOUNT DUE"; ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = d.get('total_amount', ''); ws[f'B{row}'].font = Font(bold=True, size=12)
        self._set_col_widths(ws, [30, 35])

    def _sheet_marksheet(self, wb, result):
        ws = wb.create_sheet("Marksheet")
        d = result['structured_data']
        row = self._section(ws, 1, "Student Information")
        for key in ['student_name','roll_number','class_standard','stream','school_name',
                    'board','exam_year','exam_month']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Subjects & Marks")
        self._hrow(ws, row, ['Subject', 'Max Marks', 'Marks Obtained', 'Grade', 'Pass/Fail'])
        row += 1
        for i, subj in enumerate(d.get('subjects', [])):
            fill = self.alt_fill if i % 2 == 0 else None
            for col, key in enumerate(['name','max_marks','marks_obtained','grade','pass_fail'], 1):
                c = ws.cell(row=row, column=col, value=subj.get(key, ''))
                if fill: c.fill = fill
            row += 1
        row += 1
        for key in ['total_marks','marks_obtained_total','percentage','cgpa','result','division','rank']:
            if d.get(key) is not None:
                row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        self._set_col_widths(ws, [35, 14, 18, 12, 12])

    def _sheet_ration_card(self, wb, result):
        ws = wb.create_sheet("Ration Card")
        d = result['structured_data']
        row = self._section(ws, 1, "Card Details")
        for key in ['card_type','card_number','issuing_state','issuing_department']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Head of Family")
        row = self._dict_rows(ws, row, d.get('head_of_family', {}))
        row += 1
        row = self._section(ws, row, "Family Members")
        self._hrow(ws, row, ['S.No', 'Name', 'Age', 'Relation'])
        row += 1
        for m in d.get('family_members', []):
            ws[f'A{row}'] = m.get('serial_number','')
            ws[f'B{row}'] = m.get('name','')
            ws[f'C{row}'] = m.get('age','')
            ws[f'D{row}'] = m.get('relation','')
            row += 1
        row += 1
        row = self._section(ws, row, "Ration Shop")
        row = self._dict_rows(ws, row, d.get('ration_shop', {}))
        self._set_col_widths(ws, [10, 30, 10, 20])

    # ─────────────────────────────────────────────────────────────────────────
    # Identity Documents
    # ─────────────────────────────────────────────────────────────────────────

    def _sheet_aadhar_card(self, wb, result):
        ws = wb.create_sheet("Aadhaar Card")
        d = result['structured_data']
        row = self._section(ws, 1, "Aadhaar Card Information")
        row = self._kv(ws, row, "UID Number", d.get('uid_number'))
        ws[f'B{row-1}'].font = Font(bold=True, size=13)
        row = self._kv(ws, row, "Name", d.get('name'))
        row = self._kv(ws, row, "Date of Birth", d.get('dob'))
        row = self._kv(ws, row, "Gender", d.get('gender'))
        row = self._kv(ws, row, "Relation Name", d.get('relation_name'))
        row = self._kv(ws, row, "Mobile (last 4)", d.get('mobile_linked'))
        row = self._kv(ws, row, "Issue Date", d.get('issue_date'))
        row += 1
        row = self._section(ws, row, "Address")
        row = self._dict_rows(ws, row, d.get('address', {}))
        self._set_col_widths(ws, [28, 45])

    def _sheet_pan_card(self, wb, result):
        ws = wb.create_sheet("PAN Card")
        d = result['structured_data']
        row = self._section(ws, 1, "PAN Card Information")
        row = self._kv(ws, row, "PAN Number", d.get('pan_number'))
        ws[f'B{row-1}'].font = Font(bold=True, size=13)
        row = self._kv(ws, row, "Name", d.get('name'))
        row = self._kv(ws, row, "Father's Name", d.get('father_name'))
        row = self._kv(ws, row, "Date of Birth", d.get('dob'))
        row = self._kv(ws, row, "Card Type", d.get('card_type'))
        row = self._kv(ws, row, "Issuing Authority", d.get('issuing_authority'))
        row = self._kv(ws, row, "Signature Present", d.get('signature_present'))
        self._set_col_widths(ws, [28, 45])

    def _sheet_driving_licence(self, wb, result):
        ws = wb.create_sheet("Driving Licence")
        d = result['structured_data']
        row = self._section(ws, 1, "Licence Details")
        row = self._kv(ws, row, "DL Number", d.get('dl_number'))
        ws[f'B{row-1}'].font = Font(bold=True, size=13)
        row = self._kv(ws, row, "Name", d.get('name'))
        row = self._kv(ws, row, "DOB", d.get('dob'))
        row = self._kv(ws, row, "Blood Group", d.get('blood_group'))
        row = self._kv(ws, row, "Father/Husband", d.get('father_husband_name'))
        row = self._kv(ws, row, "Issuing Authority", d.get('issuing_authority'))
        row = self._kv(ws, row, "Issue Date", d.get('issue_date'))
        row = self._kv(ws, row, "NT Validity", d.get('validity_nt'))
        row = self._kv(ws, row, "T Validity", d.get('validity_t'))
        row += 1
        row = self._section(ws, row, "Address")
        row = self._dict_rows(ws, row, d.get('address', {}))
        row += 1
        row = self._section(ws, row, "Vehicle Classes")
        self._hrow(ws, row, ['Class', 'Valid From', 'Valid To'])
        row += 1
        for vc in d.get('vehicle_classes', []):
            ws[f'A{row}'] = vc.get('class','')
            ws[f'B{row}'] = vc.get('valid_from','')
            ws[f'C{row}'] = vc.get('valid_to','')
            row += 1
        self._set_col_widths(ws, [28, 20, 20])

    def _sheet_passport(self, wb, result):
        ws = wb.create_sheet("Passport")
        d = result['structured_data']
        row = self._section(ws, 1, "Passport Details")
        row = self._kv(ws, row, "Passport Number", d.get('passport_number'))
        ws[f'B{row-1}'].font = Font(bold=True, size=13)
        for key in ['surname','given_names','nationality','dob','place_of_birth','sex',
                    'issue_date','expiry_date','place_of_issue','file_number',
                    'father_name','mother_name','spouse_name','issuing_authority']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        if d.get('address'):
            row += 1
            row = self._kv(ws, row, "Address", d.get('address'))
        row += 1
        row = self._section(ws, row, "Machine Readable Zone (MRZ)")
        ws[f'A{row}'] = d.get('mrz_line1',''); ws[f'A{row}'].font = Font(name='Courier New'); row += 1
        ws[f'A{row}'] = d.get('mrz_line2',''); ws[f'A{row}'].font = Font(name='Courier New'); row += 1
        self._set_col_widths(ws, [28, 50])

    # ─────────────────────────────────────────────────────────────────────────
    # Financial Documents
    # ─────────────────────────────────────────────────────────────────────────

    def _sheet_cheque(self, wb, result):
        ws = wb.create_sheet("Cheque")
        d = result['structured_data']
        row = self._section(ws, 1, "Cheque Details")
        row = self._kv(ws, row, "Cheque Number", d.get('cheque_number'))
        row = self._kv(ws, row, "Cheque Type", d.get('cheque_type'))
        row = self._kv(ws, row, "Date", d.get('date'))
        row = self._kv(ws, row, "Crossed", d.get('crossed'))
        row += 1
        row = self._section(ws, row, "Bank Details")
        row = self._kv(ws, row, "Bank Name", d.get('bank_name'))
        row = self._kv(ws, row, "Branch", d.get('branch_name'))
        row = self._kv(ws, row, "IFSC Code", d.get('ifsc_code'))
        row = self._kv(ws, row, "MICR Code", d.get('micr_code'))
        row = self._kv(ws, row, "Account Number", d.get('account_number'))
        row += 1
        row = self._section(ws, row, "Payment Details")
        row = self._kv(ws, row, "Payee Name", d.get('payee_name'))
        ws[f'B{row-1}'].font = Font(bold=True)
        row = self._kv(ws, row, "Amount (₹)", d.get('amount_figures'))
        ws[f'B{row-1}'].font = Font(bold=True, size=13)
        row = self._kv(ws, row, "Amount in Words", d.get('amount_words'))
        row = self._kv(ws, row, "Drawer Name", d.get('drawer_name'))
        row = self._kv(ws, row, "Memo", d.get('memo'))
        self._set_col_widths(ws, [28, 45])

    def _sheet_form_16(self, wb, result):
        ws = wb.create_sheet("Form 16")
        d = result['structured_data']
        row = self._section(ws, 1, "Certificate Details")
        row = self._kv(ws, row, "Form Part", d.get('form_part'))
        row = self._kv(ws, row, "Assessment Year", d.get('assessment_year'))
        row = self._kv(ws, row, "Certificate Number", d.get('certificate_number'))
        row += 1
        row = self._section(ws, row, "Employer")
        row = self._dict_rows(ws, row, d.get('employer', {}))
        row += 1
        row = self._section(ws, row, "Employee")
        row = self._dict_rows(ws, row, d.get('employee', {}))
        row += 1
        row = self._section(ws, row, "Income")
        row = self._dict_rows(ws, row, d.get('income', {}))
        row += 1
        row = self._section(ws, row, "Deductions (Chapter VI-A)")
        row = self._dict_rows(ws, row, d.get('deductions', {}))
        row += 1
        row = self._section(ws, row, "Tax Details")
        row = self._dict_rows(ws, row, d.get('tax', {}))
        row += 1
        if d.get('quarter_details'):
            row = self._section(ws, row, "Quarter-wise TDS")
            self._hrow(ws, row, ['Quarter', 'Amount Paid', 'TDS Deducted', 'Date of Deposit'])
            row += 1
            for q in d.get('quarter_details', []):
                ws[f'A{row}'] = q.get('quarter','')
                ws[f'B{row}'] = q.get('amount_paid','')
                ws[f'C{row}'] = q.get('tds_deducted','')
                ws[f'D{row}'] = q.get('date_of_deposit','')
                row += 1
        self._set_col_widths(ws, [32, 25, 20, 20])

    def _sheet_insurance_policy(self, wb, result):
        ws = wb.create_sheet("Insurance Policy")
        d = result['structured_data']
        row = self._section(ws, 1, "Policy Details")
        for key in ['policy_number','policy_type','insurer_name','insurer_license_number',
                    'sum_assured','premium_amount','premium_frequency',
                    'policy_term_years','premium_paying_term_years',
                    'policy_start_date','policy_end_date','next_premium_due',
                    'grace_period_days','agent_name','agent_code']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Insured Person")
        row = self._dict_rows(ws, row, d.get('insured', {}))
        row += 1
        row = self._section(ws, row, "Nominee")
        row = self._dict_rows(ws, row, d.get('nominee', {}))
        row += 1
        if d.get('riders'):
            row = self._section(ws, row, "Riders")
            self._hrow(ws, row, ['Rider Name', 'Sum Assured', 'Premium'])
            row += 1
            for r in d.get('riders', []):
                ws[f'A{row}'] = r.get('name','')
                ws[f'B{row}'] = r.get('sum_assured','')
                ws[f'C{row}'] = r.get('premium','')
                row += 1
        self._set_col_widths(ws, [32, 35, 20])

    def _sheet_gst_certificate(self, wb, result):
        ws = wb.create_sheet("GST Certificate")
        d = result['structured_data']
        row = self._section(ws, 1, "GST Registration Details")
        row = self._kv(ws, row, "GSTIN", d.get('gstin'))
        ws[f'B{row-1}'].font = Font(bold=True, size=13)
        for key in ['legal_name','trade_name','constitution','registration_date',
                    'certificate_issue_date','status','cancellation_date']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Principal Place of Business")
        row = self._dict_rows(ws, row, d.get('principal_place_of_business', {}))
        row += 1
        row = self._section(ws, row, "Jurisdiction")
        row = self._dict_rows(ws, row, d.get('jurisdiction', {}))
        row += 1
        row = self._section(ws, row, "Nature of Business")
        for nb in d.get('nature_of_business', []):
            ws[f'A{row}'] = nb; row += 1
        row += 1
        row = self._section(ws, row, "Authorized Signatory")
        row = self._dict_rows(ws, row, d.get('authorized_signatory', {}))
        self._set_col_widths(ws, [32, 45])

    # ─────────────────────────────────────────────────────────────────────────
    # Civil Records
    # ─────────────────────────────────────────────────────────────────────────

    def _sheet_birth_certificate(self, wb, result):
        ws = wb.create_sheet("Birth Certificate")
        d = result['structured_data']
        row = self._section(ws, 1, "Birth Details")
        for key in ['registration_number','child_name','gender','dob','time_of_birth',
                    'registration_date','issue_date']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Place of Birth")
        row = self._dict_rows(ws, row, d.get('place_of_birth', {}))
        row += 1
        row = self._section(ws, row, "Father's Details")
        row = self._dict_rows(ws, row, d.get('father', {}))
        row += 1
        row = self._section(ws, row, "Mother's Details")
        row = self._dict_rows(ws, row, d.get('mother', {}))
        row += 1
        row = self._section(ws, row, "Issuing Authority")
        row = self._kv(ws, row, "Authority", d.get('issuing_authority'))
        row = self._kv(ws, row, "Registrar", d.get('registrar_name'))
        row = self._kv(ws, row, "Permanent Address", d.get('permanent_address'))
        self._set_col_widths(ws, [28, 45])

    def _sheet_death_certificate(self, wb, result):
        ws = wb.create_sheet("Death Certificate")
        d = result['structured_data']
        row = self._section(ws, 1, "Death Details")
        for key in ['registration_number','deceased_name','gender','age','dod','time_of_death',
                    'cause_of_death','nationality','religion','occupation','registration_date','issue_date']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Place of Death")
        row = self._dict_rows(ws, row, d.get('place_of_death', {}))
        row += 1
        row = self._section(ws, row, "Family Details")
        row = self._kv(ws, row, "Father / Husband Name", d.get('father_husband_name'))
        row = self._kv(ws, row, "Mother Name", d.get('mother_name'))
        row = self._kv(ws, row, "Permanent Address", d.get('permanent_address'))
        row += 1
        row = self._section(ws, row, "Informant")
        row = self._kv(ws, row, "Name", d.get('informant_name'))
        row = self._kv(ws, row, "Relation", d.get('informant_relation'))
        row += 1
        row = self._section(ws, row, "Issuing Authority")
        row = self._kv(ws, row, "Authority", d.get('issuing_authority'))
        row = self._kv(ws, row, "Registrar", d.get('registrar_name'))
        self._set_col_widths(ws, [28, 45])

    def _sheet_land_record(self, wb, result):
        ws = wb.create_sheet("Land Record")
        d = result['structured_data']
        row = self._section(ws, 1, "Land Details")
        for key in ['record_type','survey_number','sub_division_number','village',
                    'taluka_tehsil','district','state','land_type','mutation_number',
                    'issue_date','remarks']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        if d.get('total_area'):
            row = self._kv(ws, row, "Total Area",
                f"{d['total_area'].get('value','')} {d['total_area'].get('unit','')}")
        row += 1
        row = self._section(ws, row, "Owners")
        self._hrow(ws, row, ['Name', "Father's Name", 'Share', 'Ownership Type', 'Address'])
        row += 1
        for own in d.get('owners', []):
            ws[f'A{row}'] = own.get('name','')
            ws[f'B{row}'] = own.get('father_name','')
            ws[f'C{row}'] = own.get('share','')
            ws[f'D{row}'] = own.get('ownership_type','')
            ws[f'E{row}'] = own.get('address','')
            row += 1
        row += 1
        row = self._section(ws, row, "Cultivation Details")
        row = self._dict_rows(ws, row, d.get('cultivation_details', {}))
        row += 1
        if d.get('encumbrances'):
            row = self._section(ws, row, "Encumbrances")
            self._hrow(ws, row, ['Type', 'Description', 'Date'])
            row += 1
            for e in d.get('encumbrances', []):
                ws[f'A{row}'] = e.get('type','')
                ws[f'B{row}'] = e.get('description','')
                ws[f'C{row}'] = e.get('date','')
                row += 1
        self._set_col_widths(ws, [28, 28, 12, 18, 35])

    def _sheet_nrega_card(self, wb, result):
        ws = wb.create_sheet("NREGA Job Card")
        d = result['structured_data']
        row = self._section(ws, 1, "Job Card Details")
        row = self._kv(ws, row, "Job Card Number", d.get('job_card_number'))
        ws[f'B{row-1}'].font = Font(bold=True, size=13)
        for key in ['village','gram_panchayat','block','district','state','pin_code','registration_date']:
            row = self._kv(ws, row, key.replace('_',' ').title(), d.get(key))
        row += 1
        row = self._section(ws, row, "Head of Household")
        row = self._dict_rows(ws, row, d.get('household_head', {}))
        row += 1
        row = self._section(ws, row, "Bank Account")
        row = self._dict_rows(ws, row, d.get('bank_account', {}))
        row += 1
        row = self._section(ws, row, "Job Seekers / Family Members")
        self._hrow(ws, row, ['Name', 'Gender', 'Age', 'Relation'])
        row += 1
        for js in d.get('job_seekers', []):
            ws[f'A{row}'] = js.get('name','')
            ws[f'B{row}'] = js.get('gender','')
            ws[f'C{row}'] = js.get('age','')
            ws[f'D{row}'] = js.get('relation','')
            row += 1
        row += 1
        row = self._section(ws, row, "Work Entries")
        self._hrow(ws, row, ['Muster No.', 'Work Name', 'From', 'To', 'Days', 'Wage Rate', 'Amount'])
        row += 1
        for we in d.get('work_entries', []):
            ws[f'A{row}'] = we.get('muster_roll_number','')
            ws[f'B{row}'] = we.get('work_name','')
            ws[f'C{row}'] = we.get('date_from','')
            ws[f'D{row}'] = we.get('date_to','')
            ws[f'E{row}'] = we.get('days_worked','')
            ws[f'F{row}'] = we.get('wage_rate','')
            ws[f'G{row}'] = we.get('amount_earned','')
            row += 1
        row += 1
        row = self._kv(ws, row, "Total Days Worked", d.get('total_days_worked'))
        row = self._kv(ws, row, "Total Wages Earned (₹)", d.get('total_wages_earned'))
        self._set_col_widths(ws, [16, 30, 14, 14, 10, 12, 14])
